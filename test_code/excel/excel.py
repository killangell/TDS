from openpyxl import Workbook

# 向sheet中写入一行数据
def insertOne(value, sheet):
    row = [value] * 4
    sheet.append(row)

# 新建excel，并创建多个sheet
if __name__ == "__main__":

    book = Workbook()
    # 新建2个自定义的sheet
    for i in range(0, 2):
        # 为每个sheet设置title，插入位置index
        sheet = book.create_sheet("sheet" + str(i + 1), i)
        # 每个sheet里设置列标题
        sheet.append(["title" + str(i + 1)] * 3)

    sheets = book.get_sheet_names()
    count = 0
    # 向sheet中插入数据
    for i in range(0, 10):
       # book["ni"] = 1
       # book["wo"] = 2
        #Sbook["ta"] = 3
        insertOne("ni", book.get_sheet_by_name(sheets[1]))
        insertOne("wo", book.get_sheet_by_name(sheets[1]))
        insertOne("ta", book.get_sheet_by_name(sheets[1]))
        count = count + 1
    book['sheet1'].cell(row=1, column=1, value="ma_quick")
    book['sheet1'].cell(row=1, column=2, value="ma_slow")
    book['sheet1'].cell(row=1, column=3, value="counter")
    book['sheet1'].cell(row=1, column=4, value="threshold")
    book['sheet1'].cell(row=1, column=5, value="profit")
    # 保存数据到.xlsx文件
    book.save("test.xlsx")
    print(str(count))