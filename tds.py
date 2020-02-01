import json
import operator

from data_source.huobi.kline import KLineDataParser
from data_source.huobi.ma import MAEx
from tds_test_huobi_kline_data import TestHuobiKline15MinData
from huobi_4hour_data_source import Huobi4HourData

# UnitTestMain.TestAll()

'''

huobi_kline_15min_json_dict = TestHuobiKline15MinData.GetTestData_15min_300()
huobi_kline_15min_json_str = json.dumps(huobi_kline_15min_json_dict)
kline_elem_list_15min = []
kline_elem_list_15min = KLineDataParser.ParseKLineDataEx(huobi_kline_15min_json_str)

kline_elem_list_result = MAEx()
kline_elem_list_result.ComputeMA(30, kline_elem_list_15min)
kline_elem_list_result.ComputeMA(60, kline_elem_list_15min)
kline_elem_list_result.PrintEqualMA(30, 60)
'''

'''
#######################################################################################################################
kline_1day_100_dict = TestHuobiKline15MinData.GetTestData_4hour_1000()
kline_1day_100_str = json.dumps(kline_1day_100_dict)
kline_elem_list_1day_100 = []
kline_elem_list_1day_100 = KLineDataParser.ParseKLineDataEx(kline_1day_100_str)

kline_elem_list_result = MAEx()
kline_elem_list_result.ComputeMA(55, kline_elem_list_1day_100)
kline_elem_list_result.ComputeMA(56, kline_elem_list_1day_100)

kline_elem_list_result.PreparePrintEqualMA()

tuple_list = []
tuple1 = kline_elem_list_result.PrintEqualMAEx(55, 56, 8)   # profit_sum = 397
tuple_list.append(tuple1)

for i in tuple_list:
    print("ma{0}, ma{1}: counter={2}, profit: {3}".format(i[0], i[1], i[2], i[3]))

#######################################################################################################################

'''

import time
import logging
from openpyxl import Workbook

def GetCurrentTimeStampStr():
    timeNum = int(round(time.time()))
    timeStamp = float(timeNum)
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    return (otherStyleTime)


logging.basicConfig(level=logging.DEBUG,  # 控制台打印的日志级别
                    filename='tds.log',
                    filemode='w',  ##模式，有w和a，w就是写模式，每次都会重新写日志，覆盖之前的日志
                    # a是追加模式，默认如果不写的话，就是追加模式
                    format= '%(asctime)s : %(message)s'
                    #'%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s'
                    # 日志格式
                    )

# Parameters

test_all = False
if test_all:
    data_source_list = [
        TestHuobiKline15MinData.GetTestData_1min_1000,
        TestHuobiKline15MinData.GetTestData_5min_1000,
        TestHuobiKline15MinData.GetTestData_15min_1000,
        TestHuobiKline15MinData.GetTestData_30min_1000,
        TestHuobiKline15MinData.GetTestData_1hour_1000,
        TestHuobiKline15MinData.GetTestData_4hour_1000,
        TestHuobiKline15MinData.GetTestData_1day_1000,
        TestHuobiKline15MinData.GetTestData_1week_1000]
    # range_quick_start = 2
    RANGE_QUICK_INIT = 2
    range_quick_end = 61
    # range_slow_start = range_quick_start + 1
    range_slow_end = 91
    range_threshold_start = 0
    range_threshold_end = 201
    log_verbose = False
else:
    data_source_list = [
        Huobi4HourData.GetTestData_4hour_500,
        Huobi4HourData.GetTestData_4hour_1000,
        Huobi4HourData.GetTestData_4hour_1500,
        Huobi4HourData.GetTestData_4hour_2000,
                        ]
    # range_quick_start = 2
    RANGE_QUICK_INIT = 2
    range_quick_end = 61
    # range_slow_start = range_quick_start + 1
    range_slow_end = 91
    range_threshold_start = 0
    range_threshold_end = 201
    log_verbose = False

data_source_index = 0
# tuple_list_all = []
for data_source in data_source_list:
    kline_data_dict = None
    kline_data_str = None
    kline_data_parse_list = None
    kline_elem_list = None
    tuple_list_from_one_data_source = []

    kline_data_dict = data_source()
    kline_data_str = json.dumps(kline_data_dict)
    kline_data_parse_list = []
    kline_data_parse_list = KLineDataParser.ParseKLineDataEx(kline_data_str)

    kline_elem_list = MAEx()
    for ma in range(2, 91):
        kline_elem_list.ComputeMA(ma, kline_data_parse_list)

    kline_elem_list.PreparePrintEqualMA()

    range_quick_start = RANGE_QUICK_INIT
    for ma_qk in range(range_quick_start, range_quick_end):
        logging.debug("{0}, ma_qk={1} start".format(GetCurrentTimeStampStr(), ma_qk))
        for ma_sl in range(range_quick_start + 1, range_slow_end):
            if ma_qk >= ma_sl:
                continue
            for thresh in range(range_threshold_start, range_threshold_end):
                tuple = None
                tuple = kline_elem_list.PrintEqualMAEx(ma_qk, ma_sl, thresh, log_verbose)
                if tuple and tuple[4] > 0:
                    tuple_list_from_one_data_source.append(tuple)
        range_quick_start += 1
        logging.debug("{0}, ma_qk={1} end".format(GetCurrentTimeStampStr(), ma_qk))
    data_source_index += 1

    tuple_list_from_one_data_source.sort(key=operator.itemgetter(4))
    print(data_source.__name__)
    # tuple_list_all.append((data_source.__name__, data_source.__name__))
    # tuple_list_all.append(tuple_list_from_one_data_source)
    # tuple_list_all.append((data_source.__name__, data_source.__name__))

    wb = Workbook()
    wb.create_sheet("Sheet")
    # create title
    wb['Sheet'].cell(row=1, column=1, value="ma_quick")
    wb['Sheet'].cell(row=1, column=2, value="ma_slow")
    wb['Sheet'].cell(row=1, column=3, value="counter")
    wb['Sheet'].cell(row=1, column=4, value="threshold")
    wb['Sheet'].cell(row=1, column=5, value="profit")
    wb['Sheet'].cell(row=1, column=6, value="positive")
    wb['Sheet'].cell(row=1, column=7, value="negative")
    wb['Sheet'].cell(row=1, column=8, value="win")
    idx = 0
    for i in tuple_list_from_one_data_source:
        if i:
            row_index = idx + 2
            wb['Sheet'].cell(row=row_index, column=1, value=i[0])
            wb['Sheet'].cell(row=row_index, column=2, value=i[1])
            wb['Sheet'].cell(row=row_index, column=3, value=i[2])
            wb['Sheet'].cell(row=row_index, column=4, value=i[3])
            wb['Sheet'].cell(row=row_index, column=5, value=i[4])
            wb['Sheet'].cell(row=row_index, column=6, value=i[5])
            wb['Sheet'].cell(row=row_index, column=7, value=i[6])
            wb['Sheet'].cell(row=row_index, column=8, value=i[7])

            # logging.debug(
            #     "idx={5} ma{0}, ma{1}: counter={2}, threshold={3}, profit={4}".format(i[0], i[1], i[2], i[3], i[4], idx))
        idx += 1
    wb.save("{0}.xlsx".format(data_source.__name__))

'''
wb = None
wb_name = "empty"
wb_name_old = None
for tuple_list_one in tuple_list_all:
    if len(tuple_list_one) == 2:
        wb_name = tuple_list_one[0]
        if wb_name != wb_name_old:
            wb = Workbook()
            wb.create_sheet("Sheet")
            # create title
            wb['Sheet'].cell(row=1, column=1, value="ma_quick")
            wb['Sheet'].cell(row=1, column=2, value="ma_slow")
            wb['Sheet'].cell(row=1, column=3, value="counter")
            wb['Sheet'].cell(row=1, column=4, value="threshold")
            wb['Sheet'].cell(row=1, column=5, value="profit")
            wb['Sheet'].cell(row=1, column=6, value="positive")
            wb['Sheet'].cell(row=1, column=7, value="negative")
            wb['Sheet'].cell(row=1, column=8, value="win")
            wb_name_old = wb_name

        # logging.debug(
        #    "func_name: ffffffffffffffffffffffffffffffffffffffffffffffffffffff: {0}, {1}".format(tuple_list_one[0], tuple_list_one[1]))
    else:
        idx = 0
        for i in tuple_list_one:
            if i:
                row_index = idx + 2
                wb['Sheet'].cell(row=row_index, column=1, value=i[0])
                wb['Sheet'].cell(row=row_index, column=2, value=i[1])
                wb['Sheet'].cell(row=row_index, column=3, value=i[2])
                wb['Sheet'].cell(row=row_index, column=4, value=i[3])
                wb['Sheet'].cell(row=row_index, column=5, value=i[4])
                wb['Sheet'].cell(row=row_index, column=6, value=i[5])
                wb['Sheet'].cell(row=row_index, column=7, value=i[6])
                wb['Sheet'].cell(row=row_index, column=8, value=i[7])

                # logging.debug(
                #     "idx={5} ma{0}, ma{1}: counter={2}, threshold={3}, profit={4}".format(i[0], i[1], i[2], i[3], i[4], idx))
            idx += 1
    wb.save("{0}.xlsx".format(wb_name))
'''